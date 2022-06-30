import sys
from PyQt5.QtWidgets import QWidget,QApplication,QMessageBox,QMainWindow,QFileDialog
from PyQt5 import uic
import pandas as pd
from openpyxl import load_workbook
import requests
import json
import datetime
import os

global in_path
global out_path

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_path, relative_path)

form = resource_path("test.ui")
form_class = uic.loadUiType(form)[0]
def get_exchangeRate(date):
    data_ = 1.0
    while data_==1.0:
        date_now = date.strftime('%Y%m%d')
        url = 'https://www.koreaexim.go.kr/site/program/financial/exchangeJSON'
        my_api_key = 'KkLgyfdes1iOSXgIcguNjRRE4KYFPql6'

        url_ = url + '?authkey=' + my_api_key + '&searchdate=' + date_now + '&data=AP01'

        result = requests.get(url_)
        datas = json.loads(result.text)
        data = 0

        for i in datas:
            if i['cur_unit'] == 'USD':
                data = i['deal_bas_r']
                data_ = data.replace(",", "")
                data_=float(data_)
        date=date-datetime.timedelta(days=1)

    return data_
#데이터프레임 >> 리스트
def dfTolist(df,dfname):
    return df[dfname].values.tolist()
#Qtwidgets의 QMainWindow, ui파일의 form_class 상속
class WindowClass(QMainWindow, form_class):
    def __init__(self):
        super().__init__()
        self.setupUi(self) #UI Setup

        #fileSelect 버튼 클릭시 selectFunction 메서드 동작   
        self.fileSelect.clicked.connect(self.selectFunction)
        self.fileSelect_2.clicked.connect(self.selectFunction2)
        #comboBox의 내용 변경시 printShtname 메서드 동작
        self.execute.clicked.connect(self.planToResult)
    #selectFunction 메서드 정의
    def selectFunction(self):

        #filePath 출력하는 부분 초기화
        self.filePath.clear()
        #comboBox 출력하는 부분 초기화
        self.comboBox.clear()
        #선택한 엑셀 파일 경로를 받아옴 : 튜플 타입으로 받아오며 0번재 요소가 주소값 string이다.
        path = QFileDialog.getOpenFileName(self, 'Open File', '', 'All File(*);; xlsx File(*.xlsx)')
        #filePath에 현재 읽어온 엑셀 파일 경로를 입력한다.(절대경로)
        self.filePath.setText(path[0])

        #위 절대 경로 활용해 openpyxl workbook 객체 생성
        wb = load_workbook(path[0])
        #설정한 workbook의 시트리스트를 읽어온다.
        self.shtlist = wb.sheetnames
        #시트리스트를 반복문으로 진행
        for sht in self.shtlist:
            #콤보박스의 addItem을 사용하여 항목 추가(addItem의 요소는 문자열 타입)
            self.comboBox.addItem(sht)

    def selectFunction2(self):

        # filePath 출력하는 부분 초기화
        self.filePath_2.clear()
        # comboBox 출력하는 부분 초기화
        self.comboBox_2.clear()
        # 선택한 엑셀 파일 경로를 받아옴 : 튜플 타입으로 받아오며 0번재 요소가 주소값 string이다.
        path = QFileDialog.getOpenFileName(self, 'Open File', '', 'All File(*);; xlsx File(*.xlsx)')
        # filePath에 현재 읽어온 엑셀 파일 경로를 입력한다.(절대경로)
        self.filePath_2.setText(path[0])

        # 위 절대 경로 활용해 openpyxl workbook 객체 생성
        wb = load_workbook(path[0])
        # 설정한 workbook의 시트리스트를 읽어온다.
        self.shtlist = wb.sheetnames

        # 시트리스트를 반복문으로 진행
        for sht in self.shtlist:
            # 콤보박스의 addItem을 사용하여 항목 추가(addItem의 요소는 문자열 타입)
            self.comboBox_2.addItem(sht)

    def planToResult(self):

        in_sheet=str(self.comboBox.currentText())
        in_df = pd.read_excel(self.filePath.text(), sheet_name=in_sheet, header=4, usecols=[0, 1, 2, 5, 8, 11],
                              names=['type', 'name', 'fmi', 'amount', 'price', 'date'])
        in_df = in_df.dropna()

        out_wb = load_workbook(self.filePath_2.text())
        out_sheet = str(self.comboBox_2.currentText())

        out_ws = out_wb[out_sheet]
        type_list = dfTolist(in_df, 'type')
        name_list = dfTolist(in_df, 'name')
        fmi_list = dfTolist(in_df, 'fmi')
        amount_list = dfTolist(in_df, 'amount')
        price_list = dfTolist(in_df, 'price')
        date_list = dfTolist(in_df, 'date')

        in_df2 = pd.read_excel(self.filePath.text(), sheet_name=in_sheet, header=4, usecols=[4, 6],
                               names=['item', 'danga'])
        in_df2 = in_df2.dropna()
        item_list = dfTolist(in_df2, 'item')
        danga_list = dfTolist(in_df2, 'danga')

        for i in range(len(type_list)):
            krw = get_exchangeRate(date_list[i])
            out_ws.cell(row=4 + i, column=5, value=krw*price_list[i])
            out_ws.cell(row=4 + i, column=1, value=date_list[i])
            out_ws.cell(row=4 + i, column=3, value=price_list[i])
            out_ws.cell(row=4 + i, column=6, value=amount_list[i])
            out_ws.cell(row=4 + i, column=7, value=fmi_list[i])
            out_ws.cell(row=4 + i, column=8, value=type_list[i])
            out_ws.cell(row=24 + i, column=1, value=date_list[i])
            out_ws.cell(row=24 + i, column=3, value=item_list[i])
            out_ws.cell(row=24 + i, column=4, value=amount_list[i])
            out_ws.cell(row=24 + i, column=5, value=danga_list[i]*amount_list[i])
            out_ws.cell(row=24 + i, column=6, value=krw*danga_list[i]*amount_list[i])
            out_ws.cell(row=24 + i, column=7, value=fmi_list[i])
            out_ws.cell(row=24 + i, column=8, value=type_list[i])





        out_wb.save(self.filePath_2.text())


#GUI 출력 부분
if __name__ == "__main__":
    app = QApplication(sys.argv)
    myWindow = WindowClass()
    myWindow.show()
    app.exec_()